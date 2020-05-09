import openpyxl

def getRowCount (file, sheetName):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumn (file, sheetName):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def readData (file, sheetName, rownum, columnno):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnno).value

def writeData (file, sheetName, rownum, columnno, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columnno).value = data
    wb.save(file)